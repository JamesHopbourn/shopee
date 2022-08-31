#!/usr/bin/env python3

import openpyxl
from sys import argv

data = {}
path = argv[1] if(len(argv) == 2) else input('拖入 Excel 文件：')
workbook = openpyxl.load_workbook(path)
worksheet = workbook.active
for i in range(2, worksheet.max_row+1):
	for j in range(1, worksheet.max_column+1):
		head = worksheet.cell(row=1, column=j).value
		item = worksheet.cell(row=i, column=j).value
		data[head] = item
	print(data)