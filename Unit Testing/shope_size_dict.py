#!/usr/bin/env python3
import openpyxl

path = '/Users/james/Code/shopee/工作簿3.xlsx'
# 开始处理数据
data = {}
result = {}
wb = openpyxl.load_workbook(path)
ws = wb.active
for i in range(2, ws.max_row+1):
	for j in range(1, ws.max_column+1):
		head = ws.cell(row=1, column=j).value
		item = ws.cell(row=i, column=j).value
		data[head] = item

sizes = {
	"EU36": "BR34=EU36=22.5CM=US5.5",
	"EU365": "BR34.5=EU36.5=23CM=US6",
	"EU37": "BR35.5=EU37=23.5CM=US6.5",
	"EU38": "BR36=EU38=24CM=US7",
	"EU385": "BR36.5=EU38.5=24.5CM=US7.5",
	"EU39": "BR37=EU39=25CM=US8",
	"EU40": "BR38=EU40=25CM=US7",
	"EU405": "BR38.5=EU40.5=25.5cm=US7.5",
	"EU41": "BR39=EU41=26CM=US8",
	"EU42": "BR40=EU42=26.5CM=US8.5",
	"EU425": "BR40.5=EU42.5=27CM=US9",
	"EU43": "BR41=EU43=27.5CM=US9.5",
	"EU44": "BR42=EU44=28CM=US10",
	"EU445": "BR42.5=EU44.5=28.5CM=US10.5",
	"EU45": "BR43=EU45=29CM=US11",
	"EUR46": "BR44=EUR46=29.5CM=US11.5"
}

result = []
text = list(data['SKUsize'].upper().split(','))
for i in range(len(text)):
	result.append(sizes[text[i]])	
print(result)