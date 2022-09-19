#!/usr/bin/env python3
import re
import os
import json
import string
import requests
import openpyxl
from sys import exit
from sys import argv
from urllib3 import *
from time import sleep
from random import choice
from random import randint
disable_warnings()

data = []
#path = input('拖入 Excel 文件：')
path = '/Users/james/Code/shopee/new.xlsx'
workbook = openpyxl.load_workbook(path)
worksheet = workbook.active
for i in range(2, worksheet.max_row+1):
	temp = {}
	for j in range(1, worksheet.max_column+1):
		head = worksheet.cell(row=1, column=j).value
		item = worksheet.cell(row=i, column=j).value
		temp[head] = item
	data.append(temp)
	
def checkProcess(i, item):
	if (not data[i][item] and not checkResult.get(item)):
		checkResult.update({item: [str(i+1)]})
	elif (not data[i][item]):
		checkResult[item].append(str(i+1))
		

# 检测文件属性值是否有效
checkResult = {}
check = ['文件夹名', '商品分类', '商品名称', '商品描述', '价格', '库存', '发货时间', '包装重量']
[checkProcess(i,item) for i in range(len(data)) for item in check]

if (len(checkResult.keys())):   
	print("以下商品属性行异常", end="\n\n")
	[print(f"{item}：第 {'、'.join(checkResult[item])} 行") for item in checkResult]
	input('\n输入回车退出程序')
	exit()
print("文件属性状态正常")