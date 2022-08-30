#!/usr/bin/env python3
import re
import os
import json
import string
import requests
import openpyxl
from sys import exit
from urllib3 import *
from time import sleep
from uuid import uuid4
from random import choice
disable_warnings()

# cookies 文件解析
def parse_cookies():
	data_cookie, header_cookies = {}, ""
	cookies = json.loads(open(os.path.join(dir_path,'cookies.txt')).read())
	for i in range(len(cookies)):
		data = json.loads(json.dumps(cookies[i]))
		data_cookie.update({data['name']: data['value']})
		header_cookies += f"{data['name']}={data['value']}; "
		header_cookies = header_cookies.replace('"', '\\"')
	return data_cookie, header_cookies

# API 通过 cookies 获取 shopID
def get_shopID():
	account_info = requests.get("https://seller.shopee.cn/api/cnsc/selleraccount/get_session/",
		cookies=data_cookie,
		auth=(),
		verify=False
	)
	status = json.loads(account_info.text)
	if ('errcode' in status):
		print(f'{status}\ncookies 失效，请重置 cookies.txt 文件\n回车退出程序')
		input()
		exit()
	return status['sub_account_info']['current_shop_id']

# 获取图片名字 调整主图顺序
def get_image_name(img_dir_name):
	images = []
	directory_name = os.path.join(dir_path, str(img_dir_name))
	for item in os.listdir(directory_name):
		if (item.endswith(('.png','.jpg','jpeg'))): images.append(item)
	images = sorted(images, key=lambda name: int(re.sub('(^.*?\(|\)\..*$)', '', name)))
	for i in range(len(images)): images[i] = f"{directory_name}/{images[i]}"
	return images

# 商品类别映射函数
def catagory_mapper():
	return {
		"篮球鞋": [100637,100726,101298],
		"跑步鞋": [100637,100726,101299],
		"训练鞋": [100637,100726,101300],
		"足球鞋": [100637,100726,101306],
		"登山鞋": [100637,100726,101305],
		"网球鞋": [100637,100726,101301],
		"排球鞋": [100637,100726,101302],
		"室内足球鞋": [100637,100726,101304],
		"女包背包": [100016, 100089],
		"女包手拿包": [100016, 100091],
		"女包腰包胸包": [100016, 100092],
		"女包托特包": [100016, 100093],
		"女包手提包": [100016, 100094],
		"女包斜挎包单肩包": [100016, 100095],
		"女包钱包其他": [100016, 100096, 100343],
		"女包钱包卡包": [100016, 100096, 100338],
		"女包钱包零钱包": [100016, 100096, 100339],
		"女包钱包手机钥匙包": [100016, 100096, 100340],
		"女包钱包双折三折钱包": [100016, 100096, 100341],
		"女包钱包长款钱包": [100016, 100096, 100342],
		"女包包包配件背带": [100016, 100097, 100344],
		"女包包包配件挂钩": [100016, 100097, 100345],
		"女包包包配件吊饰": [100016, 100097, 100346],
		"女包包包配件收纳包": [100016, 100097, 100347],
		"女包包包配件清洁保养用品": [100016, 100097, 100348]
	}[data['商品分类']]

# 鞋子品牌映射函数
def get_brandID_list():
	brand = {}
	for key in [[101298, 0],[100089, 0]]:
		flag = True
		while(flag):
			brand_list = requests.get(f"https://seller.shopee.cn/api/v3/mtsku/get_mtsku_brand_list?brand_status=1&category_ids={key[0]}&cursor={key[1]}&limit=100",
			cookies=data_cookie,
			verify=False
			)
			data = json.loads(brand_list.text)['data']['list'][0]
			flag = data.get('page_info')['has_next']
			key[1] = data.get('page_info')['cursor']
			for item in data['brand_list']:
				brand.update({item['name'].lower(): item['brand_id']})
	return brand
			
# 根据鞋码数量重复生成数据
def generate_repeat_data(size_options):
	model_list = []
	if(data[second_format] == None):
		for i in range(len(size_options)):
			model_info = {
				"mtsku_model_id": 0,
				"seller_sku": "",
				"stock_setting_list": [{"sellable_stock": data['库存']}],
				"normal_price": f"{data['价格']}",
				"is_default": False,
				"tier_index": []
			}
			model_info['tier_index'].append(i)
			model_list.append(model_info)
	else:
		for i in range(len(size_options)):
			for j in range(len(data[second_format].split(','))):
				model_info = {
					"mtsku_model_id": 0,
					"seller_sku": "",
					"stock_setting_list": [{"sellable_stock": data['库存']}],
					"normal_price": f"{data['价格']}",
					"is_default": False,
					"tier_index": []
				}
				model_info['tier_index'].extend([j, i])
				model_list.append(model_info)
	return model_list

def generate_request_data(size_options, model_list, images):
	request_data = {
		"description": data['商品描述'],
		"tier_variation": [],
		"video_list": [],
		"model_list": model_list,
		"condition": 1,
		"category_path": catagory_mapper(),
		"seller_sku": "",
		"ds_cat_rcmd_id": "",
		"description_type": "normal",
		"weight": f"{data['包装重量']}",
		"brand_id": brand[data['品牌'].lower()],
		"days_to_ship": data['发货时间'],
		"size_chart": "",
		"attributes": [],
		"dimension": {},
		"images": images[:9],
		"mtsku_item_id": 0,
		"name": data['商品名称'].strip()
	}
	attributes = [
		['子母包', 100221, {"是":1800,"否":1792}],
		['尺寸', 100218, {"微型":1793,"其他":1804}],
		['性别', 100022, {"女":652,"男":662,"男女皆宜":674}],
		['场合', 100155, {"办公":1527,"休闲":1387,"正式":1433,"其他":1397,"户外活动":1515,"运动":1502}],
		['材质', 100134, {"棉":1149,"帆布":1129,"皮革":1221,"尼龙":1165,"其他":1257,"PVC":1178,"合成皮":1280,"纺织":1285}],
		['图案', 100162, {"迷彩":1508,"点缀":1518,"刺绣":1546,"花":1447,"图案":1561,"标志":1573,"单色":1463,"波点":1474,"印花":1486,"方格/格子":1439,"条纹":1495}],
		['包包款式', 100216, {"波士顿包":1797,"水桶包":1805,"相机包":1816,"链条包":1826,"半月包":1835,"饺子包":1846,"邮差包":1855,"其他":1867,"马鞍包":1878,"小方包":1888}]
	]
	attr = [a for a, _, _ in attributes]
	for item in attributes:
		if(data.get(item[0])):
			request_data['attributes'].append({
				"attribute_id": attributes[attr.index(item[0])][1],
				"attribute_value_id": attributes[attr.index(item[0])][2][data[item[0]]]
			})
	if(data.get(second_format)): request_data['tier_variation'].append({
			"images": images[:len(data[second_format].split(','))],
			"name": f"{second_format.replace('SKU', '')}",
			"options": data[second_format].split(',')
		})
	request_data['tier_variation'].append({
		"images": [],
		"name": "size",
		"options": size_options
	})
	return request_data

# 上传商品图片
def get_image_hash(image_path):
	sets, character = [], [string.ascii_lowercase, string.ascii_uppercase, string.digits]
	[sets.append(i[j]) for i in character for j in range(len(i))]
	[character.append(choice(sets)) for i in range(16)]
	try:
		print(image_path)
		image_encode = open(image_path, 'rb').read().decode('latin-1')
		request = requests.post("https://seller.shopee.cn/api/v3/general/upload_image/",
			params={
				"SPC_CDS_VER": "2",
				"cnsc_shop_id": shopID,
			},
			data=f"""------WebKitFormBoundary{''.join(character[-16:])}
Content-Disposition: form-data; name="file"; filename="blob"
Content-Type: image/jpeg

{image_encode}
------WebKitFormBoundary{''.join(character[-16:])}--

""",
			headers={
				"Host": "seller.shopee.cn",
				"accept": "application/json, text/plain, */*",
				"accept-language": "zh-CN,zh;q=0.9",
				"content-type": f"multipart/form-data; boundary=----WebKitFormBoundary{''.join(character[-16:])}",
				"origin": "https://seller.shopee.cn",
				"sc-fe-session": f"{uuid4()}",
				"sc-fe-ver": "56878",
				"sec-ch-ua": "\".Not/A)Brand\";v=\"99\", \"Google Chrome\";v=\"103\", \"Chromium\";v=\"103\"",
				"sec-ch-ua-mobile": "?0",
				"sec-ch-ua-platform": "\"macOS\"",
				"sec-fetch-dest": "empty",
				"sec-fetch-mode": "cors",
				"sec-fetch-site": "same-origin",
				"user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"
			},
			cookies=data_cookie,
			auth=(),
			verify=False
		)
		return json.loads(request.text)['data']['resource_id']
	except (Exception):
		network_status = 1
		print(f'网络异常或者存在其他错误\n{Exception}')
		
def send_request(post_data):
	try:
		response = requests.post(
			url="https://seller.shopee.cn/api/v3/mtsku/create_mtsku/",
			params={
				"cnsc_shop_id": shopID,
			},
			headers={
				"Cookie": f"{header_cookies}",
				"Content-Type": "application/json; charset=utf-8",
			},
			data=json.dumps(post_data),
			verify=False
		)
		status = json.loads(response.text)
		if (status['code'] == -2) and ('mtsku.CreateMtskuRequest.Name' in status['message']):
			print("商品名称有误，请修改后重试")
		elif (status['code'] == 100010003 and ('mtsku title or description language is illegal ' in status['message'])):
			print('商品名称或者描述有误')
		elif (status['code'] == 0):
			print("上架成功 ✅")
		else:
			print(f"{print(status)}\n失败")
		print()
		return status['code']
	except (Exception):		
		network_status = 1
		print(f'网络异常或者存在其他错误\n{Exception}')
	
# 函数入口定义
if __name__=="__main__":
	print('虾皮自动上架软件 ver:2.0.0 2022-08-22\n软件问题反馈可联系微信：JamesHopbourn\n')
	# 定义两个 path
	path = input('拖入 Excel 文件：')
	dir_path = os.path.dirname(path)
	# 解析 cookies
	data_cookie = parse_cookies()[0]
	header_cookies = parse_cookies()[1]
	# 初始化信息
	shopID = get_shopID()
	brand = get_brandID_list()
	# 开始处理数据
	network_status = 0
	data, result = {}, {}
	wb = openpyxl.load_workbook(path)
	ws = wb.active
	for i in range(2, ws.max_row+1):
		for j in range(1, ws.max_column+1):
			head = ws.cell(row=1, column=j).value
			item = ws.cell(row=i, column=j).value
			data[head] = item
		if(data['上架情况'] == '完成' or data['暂停时间'] == None): continue
		print(f"开始上架第 {i-1} 个产品")
		# 随机数的暂停时间
		print(f"随机暂停时间：{data['暂停时间']}秒")
		sleep(data['暂停时间'])
		# 定义数组，存放图片 ID
		images = []
		for image in get_image_name(data['文件夹名']):
			images.append(get_image_hash(image))
			if network_status == 1: continue
		print(f"上传图片数量：{len(images)}")
		# 获取第二个SKU信息
		# second_format = list(filter(lambda k: k.startswith('SKU'), data))[0]
		second_format = 'SKUcolor'
		# 从 Excel 读取性别信息，返回 size_options 适配的尺码信息
		size_options = list(data['SKUsize'].split(','))
		# 根据尺码信息的数量构造库存价格对应信息
		model_list = generate_repeat_data(size_options)
		# 根据上面的两个参数构造整体的数据包
		request_data = generate_request_data(size_options, model_list, images)
		# 发送数据包上架商品返回状态码
		launch_status_code = send_request(request_data)
		# 判断网络状态
		if network_status == 1: continue
		# 根据状态码的情况决定是否追加
		if (launch_status_code == 0): result.update({i: '完成'})
# 修改 Excel 产品上架情况单元格值
index = list(data.keys()).index('上架情况')
for key in result:
	ws.cell(row=key, column=index+1, value=result[key])
wb.save(path)
print("自动上架完成")
print("回车退出程序")
input()
